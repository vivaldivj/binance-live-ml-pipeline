from binance.client import Client
import requests
import pandas as pd
from datetime import datetime, timezone, timedelta
import numpy as np 
from binance.enums import *
import sys
from openpyxl import Workbook, load_workbook
from binance.exceptions import BinanceAPIException



def colocar_orden(moneda_seleccionada, factor,cantidad_factor):
    # Reemplaza con tu API Key y Secret Key de Binance
    api_key = ""
    api_secret = ""

    # Crea un cliente Binance
    client = Client(api_key, api_secret)

    # Obtiene el saldo de tu cuenta de futuros
    account_balance = client.futures_account_balance()

    # Encuentra el saldo total en USDT (o la moneda que prefieras)
    total_balance = 0.0
    for asset in account_balance:
        if asset['asset'] == 'USDT':  # o cambia 'USDT' por la moneda que prefieras
            total_balance = float(asset['balance'])
            break

    # Imprime el saldo total
    print()
    #print(f"*   Saldo total de la cuenta de futuros: {total_balance} USDT   *")
    #print(f'*   Factor de redondeo: {factor}')
    print()


    # Verifica si ya hay una operación abierta en el mismo símbolo
    posiciones_abiertas = client.futures_position_information()

    # Variable para determinar si se puede abrir una nueva posición
    puede_abrir = True

    for posicion in posiciones_abiertas:
        # Si ya hay una posición abierta en el mismo símbolo
        if posicion['symbol'] == f"{moneda_seleccionada}USDT":
            # Verifica si la posición tiene un valor distinto de cero (es decir, está abierta)
            position_amt = float(posicion['positionAmt'])
            if position_amt != 0:
                # Si la posición es long, no se puede abrir otra operación short
                if position_amt > 0:
                    print(f"Toma nota sobre {position_amt}")
                    puede_abrir = False
                    print(f"Ya hay una operación long abierta en el símbolo {moneda_seleccionada}. No se puede abrir otra operación short.")
                    break  # Rompemos el ciclo si encontramos una posición long
                # Si la posición es short, podemos abrir otra operación short
                elif position_amt < 0:
                    puede_abrir = True  # Si la posición es short, se puede abrir otra

    # Si no se puede abrir la operación, muestra un mensaje de error
    if not puede_abrir:
        mensaje = f"No se puede abrir la operación short porque ya hay una operación long abierta en el símbolo {moneda_seleccionada}."
        print(mensaje)
        sys.exit(4)
    else:
        print(f"No hay operaciones long abiertas en el símbolo {moneda_seleccionada}. Iniciando proceso para colocación de orden short...")
        vela = bajardatosVelaEntrada(moneda_seleccionada, factor, cantidad_factor)
        calcularOrden(client, vela, total_balance, moneda_seleccionada, factor, cantidad_factor)


 
      
def bajardatosVelaEntrada(moneda_seleccionada, factor,cantidad_factor):
    # 2 Aca arrancamos con el proceso para calcular el tamaño de la opación, precio de entrada, precio de stop, etc
    # Primero nos traemos los datos de la vela de entrada, la vela que hizo el patrón Toro/Oso 180
    # Define los parámetros de la solicitud
    symbol = f'{moneda_seleccionada}USDT' ################## <-
    ##################################### <-
    interval = '30m'
    limit = 2

    # Establece la zona horaria de Ciudad de México
    tz_mexico = timezone(timedelta(hours=-6))
    # Calcula las fechas de inicio y fin en la zona horaria de Ciudad de México
    current_time = datetime.now(tz=tz_mexico)
    if current_time.minute < 30:
        end_time = current_time.replace(minute=0, second=0, microsecond=0)
    elif current_time.minute == 30 or current_time.minute == 0:
            end_time = current_time.replace(second=0, microsecond=0)
    else: 
        if current_time.minute > 30:
            end_time = current_time.replace(minute=30, second=0, microsecond=0)
    start_time = int((end_time - timedelta(minutes=30 * 2)).timestamp() * 1000)
    end_time = int(end_time.timestamp() * 1000)


    # Creamos un DataFrame vacío para almacenar todos los datos
    df_total = pd.DataFrame()

    # Realiza la solicitud a la API de Binance Futures
    url = f'https://fapi.binance.com/fapi/v1/klines?symbol={symbol}&interval={interval}&startTime={start_time}&endTime={end_time}&limit={limit}'
    response = requests.get(url)

    # Procesa los datos de la respuesta
    if response.status_code == 200:
        data = response.json()
        df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'close_time', 'quote_asset_volume', 'number_of_trades', 'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'])

        # Convierte el tiempo de UNIX a formato legible y ajusta a la zona horaria de Ciudad de México
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms', utc=True).dt.tz_convert(tz_mexico).dt.tz_localize(None)
        df['close_time'] = pd.to_datetime(df['close_time'], unit='ms', utc=True).dt.tz_convert(tz_mexico).dt.tz_localize(None)

        # Renombra las columnas
        df.columns = ['Fecha/Hora de Apertura', 'Precio de Apertura', 'Precio Máximo', 'Precio Mínimo', 'Precio de Cierre', 'Volumen', 'Fecha/Hora de Cierre', 'Volumen en Divisa de Cotización', 'Número de Operaciones', 'Volumen de Compra Base', 'Volumen de Compra Divisa de Cotización', 'Ignorar']

        # Extraer componentes de fecha y hora
        df['Mes'] = df['Fecha/Hora de Apertura'].dt.month
        df['Dia'] = df['Fecha/Hora de Apertura'].dt.day
        df['Hora'] = df['Fecha/Hora de Apertura'].dt.hour
        df['Minuto'] = df['Fecha/Hora de Apertura'].dt.minute
        df['DiaSemana'] = df['Fecha/Hora de Apertura'].dt.dayofweek + 1  # Lunes como 1, domingo como 7

        # Aqui conviene convertir los datos de las columnas open, high, low, close, volumen, a float, esto para poder hacer operaciones matemáticas después.
        df['Precio de Apertura'] = df['Precio de Apertura'].astype(float)
        df['Precio Máximo'] = df['Precio Máximo'].astype(float)
        df['Precio Mínimo'] = df['Precio Mínimo'].astype(float)
        df['Precio de Cierre'] = df['Precio de Cierre'].astype(float)
        df['Volumen'] = df['Volumen'].astype(float)

        # Concatena el DataFrame actual con el DataFrame total
        df_total = pd.concat([df_total, df], ignore_index=True)

        # Eliminamos columnas innecesarias
        df_total.drop(columns=['Ignorar', 'Volumen en Divisa de Cotización', 'Volumen de Compra Base', 'Volumen de Compra Divisa de Cotización', 'Número de Operaciones'], inplace=True)

        # Se extrae el dato de 'close_time' de la última vela extraída, es decir, de la vela número 1000, y se guarda en una variable.
        if not df.empty:
            fechaHoraUltimaVela = df['Fecha/Hora de Cierre'].iloc[-1]

        # Se actualiza start_time para que en la nueva consulta ahora empiece desde esta fecha:
        start_time = int(datetime.timestamp(fechaHoraUltimaVela) * 1000)

    # Reordenar las columnas
    df_total = df_total[['Mes', 'Dia', 'Hora', 'Minuto', 'DiaSemana', 'Fecha/Hora de Apertura', 'Precio de Apertura', 'Precio Máximo', 'Precio Mínimo', 'Precio de Cierre', 'Volumen', 'Fecha/Hora de Cierre']]
        # Eliminar la primera fila de datos
    df_total = df_total.iloc[1:]
    # Reiniciar los índices del DataFrame
    df_total.reset_index(drop=True, inplace=True)

    return df_total    
    
def calcularOrden(client,vela, total_balance, moneda_seleccionada, factor,cantidad_factor): #En esta función se calcula el riesgo de la operación, se establece precio de entrada, precio SL, precio objetivo, apalancamiento.
    apalancamiento = 30        
    precioEntrada = float(vela['Precio Mínimo'])
    precioStopLoss = float(vela['Precio Máximo'])
    
    #precioEntrada = 0.3882 #Precios arbitrarios de prueba
   # precioStopLoss = 0.3870  #Precios arbitrarios de prueba
    comision = 0.0026
    volatilidad_vela = precioStopLoss-precioEntrada 
    add_comision = precioStopLoss*comision
    add_precio =  volatilidad_vela*2
 
    poder_de_compra = ((total_balance*apalancamiento)*0.95)/precioEntrada
    riesgo = 0.03
    perdida_maxima_porOperacion = total_balance*riesgo
    tamanioVela = round((((precioStopLoss - precioEntrada) / precioEntrada)*100),2)
    tamVelaSinx100 =(precioStopLoss - precioEntrada) / precioEntrada
    if tamanioVela < 0.25:
        tamanioVela = 0.26
        tamVelaSinx100 = tamanioVela/100
    
        # Calcula la cantidad a invertir considerando el apalancamiento: "(Lo que te quieres gastar / tamVelaSinx100)/precioEntrada "
    cantidad_a_invertir = (perdida_maxima_porOperacion/tamVelaSinx100)/precioEntrada
    cantidad_a_invertir = round(cantidad_a_invertir,cantidad_factor)

    Velax2mascomision = (tamVelaSinx100*2+comision)
    
    
    
    TakeProfitprecio = round((precioEntrada - add_comision - add_precio),factor) 
    
    # Calcular precio de seguridad
    Distancia = precioEntrada - TakeProfitprecio
    ajuste = Distancia*0.69
    precio_seguridad_sinredondeo = precioEntrada - ajuste
    PrecioSeguridad = round(precio_seguridad_sinredondeo,factor)
    

    
    print(f"El precio de entrada es {precioEntrada}")
    print(f"El precio de StopLoss es {precioStopLoss}")
    print(f"Take profit precio: {TakeProfitprecio}")
    print(f"Precio de seguridad: {PrecioSeguridad}")
    print(f"El poder de compra es {poder_de_compra}")
    print(f"La vela mide {tamanioVela}")
    print(f"Vela sin ajuste {tamVelaSinx100}")
    print(f"Perdida maxima por operacion {perdida_maxima_porOperacion}")
    print(f"La cantidad a invertir es {cantidad_a_invertir}")
    print(f"*TAMAÑO OPERACION: {cantidad_a_invertir} {moneda_seleccionada}")
    
    # -------------------- CONTINUAR CON LA LOGICA PARA EJECUTAR LA ORDEN SEGUN DONDE ESTE EL PRECIO ACTUAL
    #Difiniendo si se abre la orden con Market, PostOnly o StopLimit
        # Obtener el último precio del mercado
    symbol = f'{moneda_seleccionada}USDT'  # Asegúrate de que sea el par de trading correcto

    import time
    from binance.exceptions import BinanceAPIException

    # Tu bloque de código para colocar la orden
    while True:
        try:
            precio_actual = float(client.futures_symbol_ticker(symbol=symbol)['price'])
            
            if precio_actual < precioEntrada:
                # Colocar una orden limit post only
                orden = client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_SELL,
                    type=ORDER_TYPE_LIMIT,
                    timeInForce=TIME_IN_FORCE_GTC, 
                    quantity=cantidad_a_invertir,
                    price=precioEntrada
                )
                print("Orden limit post only colocada:", orden)
            else:
                # Colocar una orden stop limit
                orden = client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_SELL,
                    type= FUTURE_ORDER_TYPE_STOP_MARKET, 
                    quantity=cantidad_a_invertir,
                    stopPrice=precioEntrada
                )
                print("Orden stop market colocada:", orden)

            # Si la orden se coloca sin error, salimos del ciclo
            break

        except BinanceAPIException as e:
            if e.code == -2021:  # Error específico de "Order would immediately trigger"
                print("Error: La orden se ejecutaría inmediatamente. Intentando nuevamente en 5 segundos...")
                time.sleep(5)  # Espera 5 segundos antes de intentar nuevamente
            else:
                # Si el error no es el esperado, lo mostramos y salimos del ciclo
                print(f"Error inesperado: {e.message}")
                break

    # Obtener el ID de la orden
    order_id = orden['orderId']
    print("La orden colocada tiene el ID:", order_id)


    from datetime import datetime, timedelta

    # Obtener la hora y minuto actual
    hora_y_minuto_actual = datetime.now().strftime('%H:%M')
    print("Hora y minuto actual:", hora_y_minuto_actual)

    # Calcular la hora y minuto límite sumando 30 minutos
    hora_actual = datetime.now() # Se obtiene para poder calulcar la hora limite
    hora_y_minuto_limite = (hora_actual + timedelta(minutes=30)).strftime('%H:%M')
    print("Hora y minuto límite para haber abierto la operación:", hora_y_minuto_limite)
    

    
    proceso_monitoreo_para_colocar_TP_SL(client,vela, precioStopLoss,PrecioSeguridad,TakeProfitprecio,cantidad_a_invertir,precioEntrada,hora_y_minuto_limite, moneda_seleccionada, factor,cantidad_factor, order_id )
           

             
import time
from datetime import datetime, timedelta

def proceso_monitoreo_para_colocar_TP_SL(client, vela, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, hora_y_minuto_limite, moneda_seleccionada, factor, cantidad_factor, order_id):
    hora_y_minuto_limite = datetime.strptime(hora_y_minuto_limite, '%H:%M').time()
    symbol = f'{moneda_seleccionada}USDT'
    stop_ajustado = False
    operacion_abierta = False
    mensaje_control1 = False
    mensaje_control2 = False
    while True:
        # Obtener la hora actual
        hora_actual = datetime.now().time()

        # Verificar si ya hay una operación abierta en {moneda_seleccionada}USDT
        open_positions = client.futures_position_information(symbol=symbol)
        position_amt = float(open_positions[0]['positionAmt'])
        
        # Verificar si la orden con el order_id ya se ejecutó
        estado_orden = client.futures_get_order(symbol=symbol, orderId=order_id)
        if estado_orden['status'] == 'FILLED':  # Orden ejecutada
            if not operacion_abierta:
                operacion_abierta = True
                print()
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(f"-   -   -   -   -   -   -   -   -   ¡Se abrió la operación! Estamos en {moneda_seleccionada}USDT. Precio de seguridad en {PrecioSeguridad}    -   -   -   -   -   -   -   -")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print()



            # Verificar si la orden con el order_id ya se ejecutó
            estado_orden = client.futures_get_order(symbol=symbol, orderId=order_id)
            if estado_orden['status'] == 'FILLED':  # Orden ejecutada
                print("No Stop Loss order. Placing Stop Loss order.")
                client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_BUY,
                    type=FUTURE_ORDER_TYPE_STOP_MARKET,
                    quantity=cantidad_a_invertir,
                    stopPrice=precioStopLoss,
                    reduceOnly=True
                )
                
                print("No Take Profit order. Placing Take Profit order (Post Only).")
                client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_BUY,
                    type=ORDER_TYPE_LIMIT,
                    timeInForce=TIME_IN_FORCE_GTC,
                    quantity=cantidad_a_invertir,
                    price=TakeProfitprecio,
                    reduceOnly=True
                )




            # Anotar en el Excel los parámetros
            resultado = 'Revisar más tarde en el celular'
            print(" *   RESULTADO GUARDADO EN EXCEL    *")
            almacenar_en_excel(vela, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, factor, cantidad_factor)
            sys.exit(0)  # Salir de la función después de procesar la operación
        
        # Si la orden no se ha ejecutado, verificar si el tiempo límite ya ha pasado
        else:
            if mensaje_control1 == False: 
                print()           
                print(f"    *   Orden colocada. Esperando a entrar en {moneda_seleccionada}USDT.")
                print()
                mensaje_control1 = True

            # Obtener el precio actual para la comparación
            ticker = client.futures_symbol_ticker(symbol=symbol)
            precio_actual = float(ticker['price'])

            # Verificar si se ha excedido el tiempo límite
            if (hora_actual >= hora_y_minuto_limite and precio_actual > precioEntrada) or (precio_actual <= PrecioSeguridad):
                print("El tiempo límite para abrir la posición se ha sido excedido. Cancelando la orden... :(")
                # Obtener la única orden abierta
                open_orders = client.futures_get_open_orders(symbol=symbol)
                if open_orders:
                    try:
                        client.futures_cancel_order(symbol=symbol, orderId=order_id)
                        print("Orden cancelada.")
                    except BinanceAPIException as e:
                        print(f"Error cancelando la orden: {e}")
                sys.exit(3)

        # Esperar 5 segundos antes de la próxima verificación
        time.sleep(5)
             
                
           
def almacenar_en_excel(vela, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, factor,cantidad_factor):
    
    mes = vela['Mes'].iloc[-1]
    dia = vela['Dia'].iloc[-1]
    hora  = vela['Hora'].iloc[-1]
    minuto = vela['Minuto'].iloc[-1]
    
    try:
        # Intenta cargar el archivo si ya existe
        wb = load_workbook('Estadisticas.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        ws = wb.active
        # Agrega los encabezados
        ws.append(['Mes', 'Día', 'Hora', 'Minuto','Moneda', 'Precio de Entrada','Precio Objetivo','Dirección', 'Resultado'])

    direccion = 'SELL'
    moneda = f'{moneda_seleccionada}'
    # Agrega una nueva fila con los datos de la operación
    ws.append([mes, dia, hora, minuto, moneda, precioEntrada,TakeProfitprecio, direccion, resultado])

    # Guarda el archivo
    wb.save('Estadisticas.xlsx')
    #print("La operación a terminado y el resultado ha sido exportado a Excel con éxito") 
                 
                
                    
    
if __name__ == "__main__":


    # Diccionario con factores correspondientes a cada criptomoneda
    factores_redondeo = {
        'BTC': 1,
        'XRP': 4,
        'ETH': 2,
        'BNB': 2,
        'ETC': 3,
        'BCH': 2,
        'DOGE': 5,
        'SOL': 2,
        'LTC': 2,
        'ADA': 4,
        'SAND': 5,
        'XLM': 5,
        'LINK': 3,
        'APT': 4           
    }

    cantidades = {
        'BTC': 3,
        'XRP': 1,
        'ETH': 3,
        'BNB': 2,
        'ETC': 2,
        'BCH': 3,
        'DOGE': 0,
        'SOL': 0,
        'LTC': 3,
        'ADA': 0,
        'SAND': 0,
        'XLM': 0,
        'LINK': 2,
        'APT': 1          
    }

    # Captura los argumentos
    moneda_seleccionada = sys.argv[1]    

    # Obtener el factor de redondeo para la criptomoneda
    factor_redondeo = factores_redondeo.get(moneda_seleccionada, None)
    factor = int(factor_redondeo) 

    # Obtener el factor para la cantidad a invertir de cada cripto
    cantidad_factor = cantidades.get(moneda_seleccionada, None)       

    colocar_orden(moneda_seleccionada, factor,cantidad_factor)