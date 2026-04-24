from binance.client import Client
import requests
import pandas as pd
from datetime import datetime, timezone, timedelta
import numpy as np 
from binance.enums import *
import sys
from openpyxl import Workbook, load_workbook
from binance.exceptions import BinanceAPIException



def colocar_orden(moneda_seleccionada, factor,cantidad_factor):
    # Reemplaza con tu API Key y Secret Key de Binance
    api_key = ""
    api_secret = ""

    # Crea un cliente Binance
    client = Client(api_key, api_secret)

    # Obtiene el saldo de tu cuenta de futuros
    account_balance = client.futures_account_balance()

    # Encuentra el saldo total en USDT (o la moneda que prefieras)
    total_balance = 0.0
    for asset in account_balance:
        if asset['asset'] == 'USDT':  # o cambia 'USDT' por la moneda que prefieras
            total_balance = float(asset['balance'])
            break

    # Imprime el saldo total
    print()
    #print(f"    * Saldo total de la cuenta de futuros: {total_balance} USDT")
    print()


    # Verifica si ya hay una operación abierta en el mismo símbolo
    posiciones_abiertas = client.futures_position_information()

    # Variable para determinar si se puede abrir una nueva posición
    puede_abrir = True

    for posicion in posiciones_abiertas:
        # Si ya hay una posición abierta en el mismo símbolo (comprobando el símbolo)
        if posicion['symbol'] == f"{moneda_seleccionada}USDT":
            # Verifica si la posición tiene un valor distinto de cero (es decir, está abierta)
            position_amt = float(posicion['positionAmt'])
            if position_amt != 0:
                # Si la posición es short, no se puede abrir otra operación long
                if position_amt < 0:
                    #print(f"Toma nota sobre {position_amt}")
                    puede_abrir = False
                    print(f"Ya hay una operación short abierta en el símbolo {moneda_seleccionada}. No se puede abrir una operación long.")
                    break  # Rompemos el ciclo si encontramos una posición short
                elif position_amt > 0:
                    puede_abrir = True  # Si no hay posiciones abiertas, se puede abrir

    # Si no se puede abrir la operación, muestra un mensaje de error
    if not puede_abrir:
        mensaje = f"No se puede abrir la operación long porque ya hay una operación short abierta en el símbolo {moneda_seleccionada}."
        print(mensaje)
        sys.exit(4)
    else:
        print(f"No hay operaciones short abiertas en el símbolo {moneda_seleccionada}. Iniciando proceso para colocación de orden long...")
        vela = bajardatosVelaEntrada(moneda_seleccionada, factor, cantidad_factor)
        calcularOrden(client, vela, total_balance, moneda_seleccionada, factor, cantidad_factor)


def bajardatosVelaEntrada(moneda_seleccionada, factor, cantidad_factor):
    # Parámetros de la solicitud
    symbol = f'{moneda_seleccionada}USDT'
    interval = '30m'
    limit = 1000

    # Establece la zona horaria de Ciudad de México
    tz_mexico = timezone(timedelta(hours=-6))
    current_time = datetime.now(tz=tz_mexico)
    if current_time.minute < 30:
        end_time = current_time.replace(minute=0, second=0, microsecond=0)
    elif current_time.minute == 30 or current_time.minute == 0:
        end_time = current_time.replace(second=0, microsecond=0)
    else: 
        if current_time.minute > 30:
            end_time = current_time.replace(minute=30, second=0, microsecond=0)
    start_time = int((end_time - timedelta(minutes=30 * 1000)).timestamp() * 1000)
    end_time = int(end_time.timestamp() * 1000)

    # DataFrame vacío
    df_total = pd.DataFrame()

    # Solicitud a la API de Binance Futures
    url = f'https://fapi.binance.com/fapi/v1/klines?symbol={symbol}&interval={interval}&startTime={start_time}&endTime={end_time}&limit={limit}'
    response = requests.get(url)

    # Procesa la respuesta de la API
    if response.status_code == 200:
        data = response.json()
        print(f"Datos descargados con éxito.")
        df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'close_time', 'quote_asset_volume', 'number_of_trades', 'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'])
        df_total = pd.concat([df_total, df], ignore_index=True)

        df_total['timestamp'] = pd.to_datetime(df_total['timestamp'], unit='ms', utc=True).dt.tz_convert(tz_mexico).dt.tz_localize(None)
        df_total['close_time'] = pd.to_datetime(df_total['close_time'], unit='ms', utc=True).dt.tz_convert(tz_mexico).dt.tz_localize(None)

        df_total.columns = ['Fecha/Hora de Apertura', 'Precio de Apertura', 'Precio Máximo', 'Precio Mínimo', 'Precio de Cierre', 'Volumen', 'Fecha/Hora de Cierre', 'Volumen en Divisa de Cotización', 'Número de Operaciones', 'Volumen de Compra Base', 'Volumen de Compra Divisa de Cotización', 'Ignorar']

        df_total['Mes'] = df_total['Fecha/Hora de Apertura'].dt.month
        df_total['Dia'] = df_total['Fecha/Hora de Apertura'].dt.day
        df_total['Hora'] = df_total['Fecha/Hora de Apertura'].dt.hour
        df_total['Minuto'] = df_total['Fecha/Hora de Apertura'].dt.minute
        df_total['DiaSemana'] = df_total['Fecha/Hora de Apertura'].dt.dayofweek + 1

        df_total['Precio de Apertura'] = df_total['Precio de Apertura'].astype(float)
        df_total['Precio Máximo'] = df_total['Precio Máximo'].astype(float)
        df_total['Precio Mínimo'] = df_total['Precio Mínimo'].astype(float)
        df_total['Precio de Cierre'] = df_total['Precio de Cierre'].astype(float)
        df_total['Volumen'] = df_total['Volumen'].astype(float)

        df_total.drop(columns=['Ignorar', 'Volumen en Divisa de Cotización', 'Volumen de Compra Base', 'Volumen de Compra Divisa de Cotización', 'Número de Operaciones'], inplace=True)

        # Aquí ya no filtramos las primeras 999 filas
        # Simplemente calculamos el ATR en todas las velas
        df_total['ATR20'] = calcular_atr(df_total, period=20)
        df_total['ATR21'] = calcular_atr(df_total, period=21)

    else:
        print(f"Error al descargar los datos: {response.status_code}")
    
    return df_total




def calcular_atr(df, period):
    """
    Calcula el Average True Range (ATR) manualmente.
    
    :param df: DataFrame con las columnas 'Precio Máximo', 'Precio Mínimo' y 'Precio de Cierre'
    :param period: Número de periodos para el cálculo del ATR
    :return: Serie de ATR calculada manualmente
    """
    high_low = df['Precio Máximo'] - df['Precio Mínimo']
    high_close = np.abs(df['Precio Máximo'] - df['Precio de Cierre'].shift(1))
    low_close = np.abs(df['Precio Mínimo'] - df['Precio de Cierre'].shift(1))

    true_range = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)

    atr = true_range.rolling(window=period).mean()  # Promedio simple del TR
    return atr




def calcularOrden(client,vela,total_balance, moneda_seleccionada, factor,cantidad_factor): #En esta función se calcula el riesgo de la operación, se establece precio de entrada, precio SL, precio objetivo, apalancamiento.
    
    apalancamiento = 30

    # Verificar si el DataFrame vela no está vacío y si los valores necesarios no son NaN
    if vela.empty:
        print("El DataFrame vela está vacío.")
        return  # Salir de la función si el DataFrame está vacío

    # Verificar que no haya NaN en las columnas 'Precio Máximo' y 'ATR20' para la última vela
    if pd.isna(vela['Precio Máximo'].iloc[-1]) or pd.isna(vela['ATR20'].iloc[-1]):
        print("Valor NaN detectado en las columnas 'Precio Máximo' o 'ATR20'.")
        return  # Salir de la función si hay NaN

    # Si todo está bien, proceder con el cálculo
    precioEntrada = float(vela['Precio Máximo'].iloc[-1])
    atr_actual = vela['ATR20'].iloc[-1]

    # Factor ATR para calcular Stop Loss y Take Profit
    factor_ATR = 3.5

    # Calcular Stop Loss (SL) y Take Profit (TP) usando ATR
    precioStopLoss = round((precioEntrada - (atr_actual * factor_ATR)), factor)
    TakeProfitprecio = precioEntrada + (atr_actual * factor_ATR * 2)  # RR = 2

    # Considerar la comisión
    comision = 0.0026
    add_comision = precioEntrada * comision

    # Ajuste de TP con la comisión
    TakeProfitprecio = round((TakeProfitprecio + add_comision), factor)

    apalancamiento = 30        

    poder_de_compra = ((total_balance*apalancamiento)*0.95)/precioEntrada
    riesgo = 0.03
    perdida_maxima_porOperacion = total_balance*riesgo
    tamanioVela = round((((precioEntrada - precioStopLoss) / precioStopLoss)*100),2)
    tamVelaSinx100 =(precioEntrada - precioStopLoss) / precioStopLoss
    if tamanioVela < 0.25:
        tamanioVela = 0.26
        tamVelaSinx100 = tamanioVela/100

        # Calcula la cantidad a invertir considerando el apalancamiento: "(Lo que te quieres gastar / tamVelaSinx100)/precioEntrada "
    cantidad_a_invertir = (perdida_maxima_porOperacion/tamVelaSinx100)/precioEntrada
    cantidad_a_invertir = round(cantidad_a_invertir,cantidad_factor)
    
    

    Velax2mascomision = (tamVelaSinx100*2+comision)
    PrecioSeguridad = round(precioEntrada*(1+(Velax2mascomision*0.69)),factor)
    

    
    print(f"El precio de entrada es {precioEntrada}")
    print(f"El precio de StopLoss es {precioStopLoss}")
    print(f"El poder de compra es {poder_de_compra}")
    print(f"La vela mide {tamanioVela}")
    print(f"Vela sin ajuste {tamVelaSinx100}")
    print(f"Perdida maxima por operacion {perdida_maxima_porOperacion}")
    print(f"La cantidad a invertir es {cantidad_a_invertir}")
    print(f"*TAMAÑO OPERACION: {cantidad_a_invertir} {moneda_seleccionada}")
    
    # -------------------- CONTINUAR CON LA LOGICA PARA EJECUTAR LA ORDEN SEGUN DONDE ESTE EL PRECIO ACTUAL
    #Difiniendo si se abre la orden con Market, PostOnly o StopLimit
        # Obtener el último precio del mercado
    symbol = f'{moneda_seleccionada}USDT'  # Asegúrate de que sea el par de trading correcto
    
    import time
    from binance.exceptions import BinanceAPIException

    # Tu bloque de código para colocar la orden
    while True:
        try:
            precio_actual = float(client.futures_symbol_ticker(symbol=symbol)['price'])
            
            if precio_actual > precioEntrada:
                # Colocar una orden limit post only
                orden = client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_BUY,
                    type=ORDER_TYPE_LIMIT,
                    timeInForce=TIME_IN_FORCE_GTC, 
                    quantity=cantidad_a_invertir,
                    price=precioEntrada
                )
                print("     * * *   *   *   *   *   *   *   *   *")
                print("     * Orden limit colocada con éxito    *")
                print("     * * *   *   *   *   *   *   *   *   *")
            else:
                # Colocar una orden stop market
                orden = client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_BUY,
                    type=FUTURE_ORDER_TYPE_STOP_MARKET, 
                    quantity=cantidad_a_invertir,
                    stopPrice=precioEntrada
                )
                print("     * * *   *   *   *   *   *   *   *   *   *")
                print("     * Orden stop market colocada con éxito  *")
                print("     * * *   *   *   *   *   *   *   *   *   *")

            # Si la orden se coloca sin error, salimos del ciclo
            break

        except BinanceAPIException as e:
            if e.code == -2021:  # Error específico de "Order would immediately trigger"
                print("Error: La orden se ejecutaría inmediatamente. Intentando nuevamente en 5 segundos...")
                time.sleep(5)  # Espera 5 segundos antes de intentar nuevamente
            else:
                # Si el error no es el esperado, lo mostramos y salimos del ciclo
                print(f"Error inesperado: {e.message}")
                break


    """ Aqui obtenemos los datos que necesitamos para monitorear la orden y colocar el TP y SL, y para hacer el registro en excel"""                
    # Obtener el ID de la orden
    order_id = orden['orderId']
    print("La orden colocada tiene el ID:", order_id)

    # Obtener los datos de la vela
    mes = vela['Mes'].iloc[-1]
    dia = vela['Dia'].iloc[-1]
    hora  = vela['Hora'].iloc[-1]
    minuto = vela['Minuto'].iloc[-1]

    direccion = 'buy'

    """Nota: Esta función unicamente se encarga de garantizar la colocación de la orden. Hasta aqui llega. A partir de aqui
       se exporta a CSV los datos necesarios para poder darle seguimiento a la orden desde el main para así, poder colocar el TP y SL cuando la orden sea FILLED. 
       El archivo CSV se deberá poder sobreescribir, ya que puede haber multiples ordenes abiertas al mismo tiempo, y la función abre una orden a la vez."""    
    crearCSV_monitorearOrden(direccion,mes,dia,hora,minuto, precioStopLoss,PrecioSeguridad,TakeProfitprecio,cantidad_a_invertir,precioEntrada, moneda_seleccionada, order_id )
    sys.exit(0)       # Hasta aqui terminaría el script actual y regresamos...

import os
import csv

def crearCSV_monitorearOrden(direccion, mes, dia, hora, minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id):
    # Ruta del archivo CSV
    ruta_csv = '4AbrirOperaciones/ordenes_monitoreo.csv'
    
    # Verificar si el archivo existe
    archivo_existe = os.path.isfile(ruta_csv)
    
    # Abrir el archivo en modo append
    with open(ruta_csv, mode='a', newline='') as file:
        writer = csv.writer(file)
        
        # Si el archivo no existe, escribir los encabezados
        if not archivo_existe:
            writer.writerow(['direccion', 'mes', 'dia', 'hora', 'minuto', 'precioStopLoss', 'PrecioSeguridad', 'TakeProfitprecio', 'cantidad_a_invertir', 'precioEntrada', 'moneda_seleccionada', 'order_id'])
        
        # Escribir los datos de la orden
        writer.writerow([direccion, mes, dia, hora, minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id])






import time
from datetime import datetime     
     
             
import time
from datetime import datetime, timedelta

def proceso_monitoreo_para_colocar_TP_SL(client, vela, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id):
    hora_y_minuto_limite = datetime.strptime(hora_y_minuto_limite, '%H:%M').time()
    symbol = f'{moneda_seleccionada}USDT'
    stop_ajustado = False
    operacion_abierta = False
    mensaje_control1 = False
    mensaje_control2 = False
    while True:
        # Obtener la hora actual
        hora_actual = datetime.now().time()

        # Verificar si ya hay una operación abierta en {moneda_seleccionada}USDT
        open_positions = client.futures_position_information(symbol=symbol)
        position_amt = float(open_positions[0]['positionAmt'])
        
        # Verificar si la orden con el order_id ya se ejecutó
        estado_orden = client.futures_get_order(symbol=symbol, orderId=order_id)
        if estado_orden['status'] == 'FILLED':  # Orden ejecutada
            if not operacion_abierta:
                operacion_abierta = True
                print()
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(f"-   -   -   -   -   -   -   -   -   ¡Se abrió la operación! Estamos en {moneda_seleccionada}USDT. Precio de seguridad en {PrecioSeguridad}    -   -   -   -   -   -   -   -")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print(" ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
                print()



            # Verificar si la orden con el order_id ya se ejecutó
            estado_orden = client.futures_get_order(symbol=symbol, orderId=order_id)
            if estado_orden['status'] == 'FILLED':  # Orden ejecutada
                print("No Stop Loss order. Placing Stop Loss order.")
                client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_SELL,
                    type=FUTURE_ORDER_TYPE_STOP_MARKET,
                    quantity=cantidad_a_invertir,
                    stopPrice=precioStopLoss,
                    reduceOnly=True
                )

                print("No Take Profit order. Placing Take Profit order (Post Only).")
                client.futures_create_order(
                    symbol=symbol,
                    side=SIDE_SELL,
                    type=ORDER_TYPE_LIMIT,
                    timeInForce=TIME_IN_FORCE_GTC,
                    quantity=cantidad_a_invertir,
                    price=TakeProfitprecio,
                    reduceOnly=True
                )

            # Anotar en el Excel los parámetros
            resultado = 'Revisar más tarde en el celular'
            print(" *   RESULTADO GUARDADO EN EXCEL    *")
            almacenar_en_excel(vela, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada)
            sys.exit(0)  # Salir de la función después de procesar la operación
        
        # Si la orden no se ha ejecutado, verificar si el tiempo límite ya ha pasado
        else:
            if mensaje_control1 == False: 
                print()           
                print(f"    *   Orden colocada. Esperando a entrar en {moneda_seleccionada}USDT.")
                print()
                mensaje_control1 = True

            # Obtener el precio actual para la comparación
            ticker = client.futures_symbol_ticker(symbol=symbol)
            precio_actual = float(ticker['price'])

            # Verificar si se ha excedido el tiempo límite para abrir la posición
            if (hora_actual >= hora_y_minuto_limite and precio_actual < precioEntrada) or (precio_actual >= PrecioSeguridad):
                print("El tiempo límite para abrir la posición se ha sido excedido. Cancelando la orden... :(")
                # Obtener la única orden abierta
                open_orders = client.futures_get_open_orders(symbol=symbol)
                if open_orders:
                    try:
                        client.futures_cancel_order(symbol=symbol, orderId=order_id)
                        print("Orden cancelada.")
                    except BinanceAPIException as e:
                        print(f"Error cancelando la orden: {e}")
                sys.exit(3)  # Salir si la orden fue cancelada

        # Esperar 5 segundos antes de la próxima verificación
        time.sleep(5)

               
                
           
def almacenar_en_excel(vela, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada):
    
    mes = vela['Mes'].iloc[-1]
    dia = vela['Dia'].iloc[-1]
    hora  = vela['Hora'].iloc[-1]
    minuto = vela['Minuto'].iloc[-1]
    
    try:
        # Intenta cargar el archivo si ya existe
        wb = load_workbook('Estadisticas.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        ws = wb.active
        # Agrega los encabezados
        ws.append(['Mes', 'Día', 'Hora', 'Minuto','Moneda', 'Precio de Entrada','Precio Objetivo','Dirección', 'Resultado'])

    direccion = 'buy'
    moneda = f'{moneda_seleccionada}'
    # Agrega una nueva fila con los datos de la operación
    ws.append([mes, dia, hora, minuto, moneda, precioEntrada,TakeProfitprecio, direccion, resultado])

    # Guarda el archivo
    wb.save('Estadisticas.xlsx')
    #print("La operación a terminado y el resultado ha sido exportado a Excel con éxito") 
                 
import os
import csv

def probar_ruta():
    # Ruta del archivo CSV de prueba
    ruta_csv = '4AbrirOperaciones/ordenes_monitoreo.csv'
    
    # Verificar si el directorio existe, si no, crearlo
    directorio = os.path.dirname(ruta_csv)
    print(f"Directorio: {directorio}")
    if not os.path.exists(directorio):
        print("Directorio no existe, creando...")
        os.makedirs(directorio)
    else:
        print("Directorio ya existe.")
    
    # Crear el archivo CSV de prueba
    with open(ruta_csv, mode='w', newline='') as file:
        writer = csv.writer(file)
        
        # Escribir los encabezados de prueba
        writer.writerow(['direccion', 'mes', 'dia', 'hora', 'minuto', 'precioStopLoss', 'PrecioSeguridad', 'TakeProfitprecio', 'cantidad_a_invertir', 'precioEntrada', 'moneda_seleccionada', 'order_id'])
        
        # Escribir una fila de datos de prueba
        writer.writerow(['buy', 1, 1, 0, 0, 0.0, 0.0, 0.0, 0.0, 0.0, 'BTC', 123456])
    
    print(f"Archivo de prueba creado en: {ruta_csv}")


                              
    
if __name__ == "__main__":


    # Diccionario con factores correspondientes a cada criptomoneda
    factores_redondeo = {
        'BTC': 1,
        'XRP': 4,
        'ETH': 2,
        'BNB': 2,
        'ETC': 3,
        'BCH': 2,
        'DOGE': 5,
        'SOL': 2,
        'LTC': 2,
        'ADA': 4,
        'SAND': 5,
        'XLM': 5,
        'LINK': 3,
        'APT': 4           
    }

    cantidades = {
        'BTC': 3,
        'XRP': 1,
        'ETH': 3,
        'BNB': 2,
        'ETC': 2,
        'BCH': 3,
        'DOGE': 0,
        'SOL': 0,
        'LTC': 3,
        'ADA': 0,
        'SAND': 0,
        'XLM': 0,
        'LINK': 2,
        'APT': 1          
    }

    # Captura los argumentos
    moneda_seleccionada = sys.argv[1]    

    # Obtener el factor de redondeo para la criptomoneda
    factor_redondeo = factores_redondeo.get(moneda_seleccionada, None)
    factor = int(factor_redondeo) 

    # Obtener el factor para la cantidad a invertir de cada cripto
    cantidad_factor = cantidades.get(moneda_seleccionada, None)       

    colocar_orden(moneda_seleccionada, factor,cantidad_factor)
    #probar_ruta()
