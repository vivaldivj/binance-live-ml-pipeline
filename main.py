import time
from binance.client import Client
import tkinter as tk
from tkinter import messagebox
import requests
import pandas as pd
from datetime import datetime, timezone, timedelta
import numpy as np 
from binance.enums import *
import subprocess
import os
from datetime import datetime
from binance.client import Client
from openpyxl import Workbook, load_workbook

def monitorear_ordenes(api_key, api_secret):
    #print("Entramos a la función")

    try:
        # Crear un objeto tipo cliente de Binance:
        client = Client(api_key, api_secret)

        # Ruta del archivo CSV
        ruta_csv = '4AbrirOperaciones/ordenes_monitoreo.csv'

        # Verificar si el archivo existe
        if not os.path.isfile(ruta_csv):
            #print("El archivo no existe")
            #print("No hay órdenes para monitorear.")
            return


        # Leer el archivo CSV
        df = pd.read_csv(ruta_csv)

        # Iterar sobre cada fila del archivo
        for index, row in df.iterrows():
            order_id = int(row['order_id'])  # Forzamos a convertirlo a int
            #print(f"La orden es {order_id}")
            direccion = row['direccion']
            precioStopLoss = row['precioStopLoss']
            PrecioSeguridad = row['PrecioSeguridad']
            TakeProfitprecio = row['TakeProfitprecio']
            cantidad_a_invertir = row['cantidad_a_invertir']
            precioEntrada = row['precioEntrada']
            moneda_seleccionada = row['moneda_seleccionada']
            moneda_seleccionada = f"{moneda_seleccionada}USDT"
            # Para el excel de las estadisticas tomamos estos datos para enviarlos a la función almacenar_en_excel
            Mes = row['mes']
            Dia = row['dia']
            Hora = row['hora']
            Minuto = row['minuto']
            # Verificar el estado de la orden
            estado_orden = client.futures_get_order(symbol=moneda_seleccionada, orderId=order_id)
            #print(f"La orden pendiente {order_id} de {moneda_seleccionada} aun no se ha llenado")
            # Si el estado de la orden es 'filled'
            if estado_orden['status'] == 'FILLED':
                if direccion == 'buy':
                    buy_colocar_stoploss_takeprofit_HEDGE(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion)
                elif direccion == 'SELL':
                    SELL_colocar_stoploss_takeprofit_HEDGE(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion)

                # Eliminar la fila del archivo CSV
                df.drop(index, inplace=True)
        # Guardar el archivo CSV actualizado
        df.to_csv(ruta_csv, index=False)

    except Exception as e:
        print(f"Error en monitorear_ordenes: {e}. Se intentará nuevamente en el siguiente ciclo.")

# Ejemplo de funciones buy_colocar_stoploss_takeprofit y SELL_colocar_stoploss_takeprofit
def buy_colocar_stoploss_takeprofit(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion):
        # Verificar si la orden con el order_id ya se ejecutó
        estado_orden = client.futures_get_order(symbol=moneda_seleccionada, orderId=order_id)
        if estado_orden['status'] == 'FILLED':  # Orden ejecutada
            print("No Stop Loss order. Placing Stop Loss order.")
            client.futures_create_order(
                symbol=moneda_seleccionada,
                side=SIDE_SELL,
                type=FUTURE_ORDER_TYPE_STOP_MARKET,
                quantity=cantidad_a_invertir,
                stopPrice=precioStopLoss,
                reduceOnly=True
            )

            print("No Take Profit order. Placing Take Profit order (Post Only).")
            client.futures_create_order(
                symbol=moneda_seleccionada,
                side=SIDE_SELL,
                type=ORDER_TYPE_LIMIT,
                timeInForce=TIME_IN_FORCE_GTC,
                quantity=cantidad_a_invertir,
                price=TakeProfitprecio,
                reduceOnly=True
            )

        # Anotar en el Excel los parámetros
        resultado = 'Revisar'
        print(" *   OPERACIÓN GUARDADA EN EXCEL    *")
        almacenar_en_excel(Mes, Dia, Hora, Minuto, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, direccion)



# Ejemplo de funciones buy_colocar_stoploss_takeprofit y SELL_colocar_stoploss_takeprofit
def buy_colocar_stoploss_takeprofit_HEDGE(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion):
    # Verificar si la orden con el order_id ya se ejecutó
    estado_orden = client.futures_get_order(symbol=moneda_seleccionada, orderId=order_id)
    
    if estado_orden['status'] == 'FILLED':  # Orden ejecutada
        print("No Stop Loss order. Placing Stop Loss order.")
        
        # Orden de Stop Loss en modo Hedge (LONG)
        client.futures_create_order(
            symbol=moneda_seleccionada,
            side=SIDE_SELL,
            positionSide='LONG',  # Asegúrate de especificar el positionSide correcto
            type=FUTURE_ORDER_TYPE_STOP_MARKET,
            quantity=cantidad_a_invertir,
            stopPrice=precioStopLoss,
            closePosition='true'
        )
        
        print("No Take Profit order. Placing Take Profit order (Post Only).")
        
        # Orden de Take Profit en modo Hedge (LONG)
        client.futures_create_order(
            symbol=moneda_seleccionada,
            side=SIDE_SELL, # "Voy a vender para cerrar"
            positionSide='LONG',  # "Tengo LONG abierto"
            type=ORDER_TYPE_LIMIT,
            timeInForce=TIME_IN_FORCE_GTC,
            quantity=cantidad_a_invertir,
            price=TakeProfitprecio,
            
        )

        # Anotar en el Excel los parámetros
        resultado = 'Revisar'
        print(" *   OPERACIÓN GUARDADA EN EXCEL    *")
        almacenar_en_excel(Mes, Dia, Hora, Minuto, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, direccion)



def SELL_colocar_stoploss_takeprofit(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion):
        # Verificar si la orden con el order_id ya se ejecutó
        estado_orden = client.futures_get_order(symbol=moneda_seleccionada, orderId=order_id)
        if estado_orden['status'] == 'FILLED':  # Orden ejecutada
            print("No Stop Loss order. Placing Stop Loss order.")
            client.futures_create_order(
                symbol=moneda_seleccionada,
                side=SIDE_BUY,
                type=FUTURE_ORDER_TYPE_STOP_MARKET,
                quantity=cantidad_a_invertir,
                stopPrice=precioStopLoss,
                reduceOnly=True
            )
            
            print("No Take Profit order. Placing Take Profit order (Post Only).")
            client.futures_create_order(
                symbol=moneda_seleccionada,
                side=SIDE_BUY,
                type=ORDER_TYPE_LIMIT,
                timeInForce=TIME_IN_FORCE_GTC,
                quantity=cantidad_a_invertir,
                price=TakeProfitprecio,
                reduceOnly=True
            )

        # Anotar en el Excel los parámetros
        resultado = 'Revisar'
        print(" *   OPERACIÓN GUARDADA EN EXCEL    *")
        almacenar_en_excel(Mes, Dia, Hora, Minuto, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, direccion)



def SELL_colocar_stoploss_takeprofit_HEDGE(client, Mes, Dia, Hora, Minuto, precioStopLoss, PrecioSeguridad, TakeProfitprecio, cantidad_a_invertir, precioEntrada, moneda_seleccionada, order_id, direccion):
    # Verificar si la orden con el order_id ya se ejecutó
    estado_orden = client.futures_get_order(symbol=moneda_seleccionada, orderId=order_id)
    
    if estado_orden['status'] == 'FILLED':  # Orden ejecutada
        print("No Stop Loss order. Placing Stop Loss order.")
        
        # Orden de Stop Loss en modo Hedge (SHORT)
        client.futures_create_order(
            symbol=moneda_seleccionada,
            side=SIDE_BUY,  # SIDE_BUY para cerrar la posición SHORT
            positionSide='SHORT',  # Especificamos la posición SHORT
            type=FUTURE_ORDER_TYPE_STOP_MARKET,
            quantity=cantidad_a_invertir,
            stopPrice=precioStopLoss,
            closePosition='true'
        )

        print("No Take Profit order. Placing Take Profit order (Post Only).")
        
        # Orden de Take Profit en modo Hedge (para reducir una posición SHORT)
        client.futures_create_order(
            symbol=moneda_seleccionada,
            side=SIDE_BUY,  # "Voy a comprar"
            positionSide='SHORT',  # "Tengo SHORT abierto"
            type=ORDER_TYPE_LIMIT,  # Tipo de orden LIMIT para el Take Profit
            timeInForce=TIME_IN_FORCE_GTC,  # La orden se mantiene hasta que se ejecute o sea cancelada
            quantity=cantidad_a_invertir,
            price=TakeProfitprecio,  # El precio al que deseas que se ejecute el Take Profit
            
        )


        # Anotar en el Excel los parámetros
        resultado = 'Revisar'
        print(" *   OPERACIÓN GUARDADA EN EXCEL    *")
        almacenar_en_excel(Mes, Dia, Hora, Minuto, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, direccion)


import os
from openpyxl import load_workbook, Workbook

def almacenar_en_excel(Mes, Dia, Hora, Minuto, precioEntrada, TakeProfitprecio, resultado, moneda_seleccionada, direccion):
    # Verificar el sistema operativo y obtener la ruta del escritorio
    home = os.environ['HOME']
    
    # Intentar las dos rutas posibles para el escritorio
    posibles_rutas = [
        os.path.join(home, 'Desktop'),      # Ruta estándar
        os.path.join(home, 'Escritorio')   # Ruta alternativa (por idioma o configuración)
    ]
    
    # Buscar la ruta que existe
    escritorio = None
    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            escritorio = ruta
            break
    
    # Si no se encuentra el escritorio, notificar al usuario
    if escritorio is None:
        print(f"Error: No se encontró el directorio de escritorio en {home}.")
        return
    
    ruta_excel = os.path.join(escritorio, 'Estadisticas.xlsx')
    try:
        # Intenta cargar el archivo si ya existe
        wb = load_workbook(ruta_excel)
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        ws = wb.active
        # Agrega los encabezados
        ws.append(['Mes', 'Día', 'Hora', 'Minuto', 'Moneda', 'Precio de Entrada', 'Precio Objetivo', 'Dirección', 'Resultado'])

    moneda = f'{moneda_seleccionada}'
    # Agrega una nueva fila con los datos de la operación
    ws.append([Mes, Dia, Hora, Minuto, moneda, precioEntrada, TakeProfitprecio, direccion, resultado])

    # Guarda el archivo en el escritorio
    wb.save(ruta_excel)
    print("La operación ha terminado y el resultado ha sido exportado a Excel con éxito en el escritorio")




def cancelar_ordenes_pendientes(api_key, api_secret):
    """
    Obtiene todas las órdenes abiertas en Binance Futures y cancela aquellas que no son Reduce-Only.

    :param api_key: Clave de API de Binance
    :param api_secret: Secreto de API de Binance
    """
    # Inicializar el cliente de Binance
    client = Client(api_key, api_secret)

    try:
        # Obtener todas las órdenes abiertas del usuario
        ordenes_abiertas = client.futures_get_open_orders()

        # Filtrar solo las órdenes que NO son Reduce-Only (es decir, órdenes para abrir posiciones)
        ordenes_pendientes = [orden for orden in ordenes_abiertas if not orden.get("reduceOnly", False)]

        print(f"\n🔹 Órdenes pendientes encontradas: {len(ordenes_pendientes)}")

        # Si no hay órdenes pendientes, mostrar mensaje
        if not ordenes_pendientes:
            print("✅ No hay órdenes pendientes para cancelar.")
            # Agregar aqui la logica para eliminar el archivo csv 
            ruta_csv = '4AbrirOperaciones/ordenes_monitoreo.csv'
            if os.path.exists(ruta_csv):
                os.remove(ruta_csv)
                print(f"Archivo {ruta_csv} eliminado exitosamente.")
            else:
                print(f"El archivo {ruta_csv} no existe.")            
            return

        # Cancelar todas las órdenes pendientes
        for orden in ordenes_pendientes:
            symbol = orden["symbol"]  # Par de la orden (ej. BTCUSDT)
            order_id = orden["orderId"]
            print(f"Se identifico orden en: - 🆔 ID: {order_id}, Cripto: {symbol}")
            # Intentar cancelar la orden
            client.futures_cancel_order(symbol=symbol, orderId=order_id)
            print(f"🚫 Orden cancelada - 🆔 ID: {order_id}, Cripto: {symbol}")
        # Agregar aqui la logica para eliminar el archivo csv 

        ruta_csv = '4AbrirOperaciones/ordenes_monitoreo.csv'
        if os.path.exists(ruta_csv):
            os.remove(ruta_csv)
            print(f"Archivo {ruta_csv} eliminado exitosamente.")
        else:
            print(f"El archivo {ruta_csv} no existe.")
    except Exception as e:
        print(f"❌ Error al cancelar órdenes: {e}")





def ejecutar_script(moneda_seleccionada,factor):
    # Definir la ruta del script
    ruta_script1 = os.path.join(BASE_DIR, '1Parte', 'ObtenerDatos.py')
    
    try:
        # Ejecutar el script y pasar las variables moneda_seleccionada y factor como argumentos
        result_script1 = subprocess.Popen(
            [sys.executable, ruta_script1, moneda_seleccionada, str(factor)]
        )
        # Esperar a que el script termine de ejecutarse
        result_script1.wait()
        if result_script1.returncode == 0:
            return 0
        elif result_script1.returncode == 1:
            return 1  
    except FileNotFoundError:
        print("El archivo correspondiente a la moneda seleccionada no se encontró.")
        return 1


def iniciarproceso(now):

    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -    -    -   -   -   -   -   -   -   -   -   -   -   -     -       -   -")
    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   INICIA PROCESO -    -   -   -   -   -   -   -   -   -   -   -   -   -")  
    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -    -    -   -   -   -   -   -   -   -   -   -   -   -     -       -   -")            
    print(f"-   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -                Fecha y hora actual: {now.strftime('%Y-%m-%d %H:%M:%S')}") 
    print()       
    # Lista de monedas y acciones
    # Lista de las 10 principales criptomonedas
    #criptomonedas = ['ADA', 'XRP', 'BCH', 'LINK','ETH','SAND','XLM', 'LTC', 'ETC', 'DOGE', 'APT' ,'BNB' ]#'SOL', 'BTC'
    criptomonedas =[ 'BNB']

    # Diccionario con factores correspondientes a cada criptomoneda
    factores = {
        'BTC': 1,
        'XRP': 10000,
        'ETH': 10,
        'BNB': 100,
        'ETC': 100,
        'BCH': 100,
        'DOGE': 100000,
        'SOL': 100,
        'LTC': 100,
        'ADA': 10000,
        'SAND': 10000,
        'XLM': 100000,
        'APT': 1000,
        'LINK': 1000
    }

    for moneda in criptomonedas:
        print()
        print(" *   ")
        print(f"    * Ejecutando para  - {moneda}")
        # Se obtiene el factor correspondiente de la criptomoneda
        factor = factores[moneda]
        
        print()        
        control_proceso = ejecutar_script(moneda,factor)
        if control_proceso == 1:
            return

def iniciarproceso_ini(now):

    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -    -    -   -   -   -   -   -   -   -   -   -   -   -     -       -   -")
    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   INICIA PROCESO -    -   -   -   -   -   -   -   -   -   -   -   -   -")  
    print(" -   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -    -    -   -   -   -   -   -   -   -   -   -   -   -     -       -   -")            
    print(f"-   -   -   -   -   -   -   -   -   -   -   -   -   -   -   -                Fecha y hora actual: {now.strftime('%Y-%m-%d %H:%M:%S')}")  
    print()    
    # Lista de monedas y acciones
    # Lista de las 10 principales criptomonedas
    #criptomonedas = ['ADA', 'XRP', 'BCH', 'LINK','ETH','SAND','XLM', 'LTC', 'ETC', 'DOGE', 'APT' ,'BNB' ]#'SOL', 'BTC'
    criptomonedas =[ 'BNB']

    # Diccionario con factores correspondientes a cada criptomoneda
    factores = {
        'BTC': 1,
        'XRP': 10000,
        'ETH': 10,
        'BNB': 100,
        'ETC': 100,
        'BCH': 100,
        'DOGE': 100000,
        'SOL': 100,
        'LTC': 100,
        'ADA': 10000,
        'SAND': 10000,
        'XLM': 100000,
        'APT': 1000,
        'LINK': 1000
    }

    for moneda in criptomonedas:
        print()
        print(" *   ")
        print(f"    * Ejecutando para  - {moneda}")
        # Se obtiene el factor correspondiente de la criptomoneda
        factor = factores[moneda]
        
        print()        
        control_proceso = ejecutar_script(moneda,factor)
        if control_proceso == 1:
            return


import os
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def crear_excel_en_escritorio():
    # Obtener la ruta del directorio Home
    home = os.environ['HOME']
    
    # Intentar las dos rutas posibles
    posibles_rutas = [
        os.path.join(home, 'Desktop'),      # Ruta estándar
        os.path.join(home, 'Escritorio')   # Ruta alternativa (por idioma o configuración)
    ]
    
    # Buscar la ruta que existe
    escritorio = None
    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            escritorio = ruta
            break
    
    # Si no se encuentra el escritorio, notificar al usuario
    if escritorio is None:
        print(f"Error: No se encontró el directorio de escritorio en {home}.")
        return
    
    # Definir la ruta del archivo Excel
    ruta_excel = os.path.join(escritorio, 'archivo_de_prueba.xlsx')
    
    try:
        # Crear un nuevo libro de trabajo y hoja de trabajo
        wb = Workbook()
        ws = wb.active
        
        # Agregar algunos datos de ejemplo
        ws.append(['Nombre', 'Edad', 'Ciudad'])
        ws.append(['Juan', 28, 'Madrid'])
        ws.append(['Ana', 22, 'Barcelona'])
        ws.append(['Luis', 35, 'Sevilla'])
        
        # Guardar el archivo en el escritorio
        wb.save(ruta_excel)
        
        print(f"El archivo Excel se ha creado exitosamente en: {ruta_excel}")
    
    except PermissionError:
        print(f"Error: No tienes permisos para escribir en {escritorio}.")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")







if __name__ == "__main__":
    # Configura tus credenciales
    api_key = ""
    api_secret = ""  
    now = datetime.now()  
    iniciarproceso_ini(now)
    while True:
        now = datetime.now()
         #Aqui vendría la función que lee el csv, y coloca sl y tp
        monitorear_ordenes(api_key, api_secret) # Esta función monitorea repetitivamente las órdenes que no se han llenado y coloca SL y TP en cuanto detecte que sean FILLED
        minutos = now.minute
        segundos = now.second
        # Si estamos entre 29:30 - 30:00 o 59:30 - 60:00, ejecutamos la función
        if (minutos == 29 and segundos >= 52) or (minutos == 59 and segundos >= 52):
            #cancelar_ordenes_pendientes(api_key, api_secret)
            cancelar_ordenes_pendientes(api_key, api_secret)
        if now.minute == 0 or now.minute == 30:     
            time.sleep(2)     
            iniciarproceso(now)
            # Aqui verificamos si al terminar de iniciarproceso() seguimos en el minuto 0 o 30, si si, dormimos 59 segundos para evitar que este imprimiendo varias veces
            now = datetime.now()
            if now.minute == 0 or now.minute == 30:
                time.sleep(59)
        else:
            # Dormir unos segundos para evitar chequear constantemente
            time.sleep(2)         
    #crear_excel_en_escritorio()            
